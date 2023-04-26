import os
import openpyxl


class readdata:



    def fetch_camapaigns_enviroment(strcampaignname):

        # Fetch campaigns and devices list whose execute status are marked as "Yes"

        # current directory
        current_dir = os.getcwd()
        parent_dir = os.path.abspath(os.path.join(current_dir, os.pardir))
        #print(parent_dir)
        # test data file path
        #test_data_file_path = parent_dir + "\\" + "testData" + "\\Test_Data.xlsx"

        test_data_file_path = "C:\\Users\\RudreshaC\\PycharmProjects\\RantCell_V1\\testData\\Test_Data.xlsx"

        campaignwb = openpyxl.load_workbook(test_data_file_path)
        campaignsheet = campaignwb["CAMPAIGNS_TOTEST"]
        campaignsheetrows = campaignsheet.max_row
        campaigns_test = []
        for r in range(2, campaignsheetrows + 1):
            values = ''
            if campaignsheet.cell(row=r, column=3).value == "Yes":
                for c in range(1, 3):
                    value = campaignsheet.cell(row=r, column=c).value.strip()
                    values = values + value + ","
                temp = values.split(",")
                campaigns_test.append(tuple(temp))
        # print(campaigns_test)

        # Fetch environment details to carry out execution based on user input

        environment_test = []
        environmentwb = openpyxl.load_workbook(test_data_file_path)
        environmentsheet = environmentwb["ENVIRONMENTS_USERINPUT_LOGIN"]
        campaignsheetrows = environmentsheet.max_row
        for r in range(2, campaignsheetrows + 1):
            values = ''
            if environmentsheet.cell(row=r, column=5).value == "Yes":
                for c in range(1, 5):
                    value = environmentsheet.cell(row=r, column=c).value.strip()
                    values = values + value + ","
                temp = values.split(",")
                environment_test.append(tuple(temp))
        # print(environment_test)

        campaign_environment = []
        for x in campaigns_test:
            va = ''
            for y in environment_test:
                va = x[0] + "," + x[1] + "," + y[0] + "," + y[1] + "," + y[2] + "," + y[3]
                temp = va.split(",")
                campaign_environment.append(tuple(temp))

        return campaign_environment

    def fetch_components(strcampaignname):

        # Fetch components list based on campaign name

        # current directory
        current_dir = os.getcwd()
        parent_dir = os.path.abspath(os.path.join(current_dir, os.pardir))
        # test data file path
        #test_data_file_path = parent_dir + "\\" + "testData" + "\\Test_Data.xlsx"
        test_data_file_path = "C:\\Users\\RudreshaC\\PycharmProjects\\RantCell_V1\\testData\\Test_Data.xlsx"
        component_list = []
        componentwb = openpyxl.load_workbook(test_data_file_path)
        componentsheet = componentwb["TC"]
        componentsheetrows = componentsheet.max_row
        componentsheetcoloumns = componentsheet.max_column
        for r in range(2, componentsheetrows + 1):
            if componentsheet.cell(row=r, column=4).value.strip() == strcampaignname:
                for c in range(5, componentsheetcoloumns + 1):
                    if componentsheet.cell(row=r, column=c).value == "Yes":
                        value = componentsheet.cell(row=1, column=c).value.replace(" ", "")
                        component_list.append(value)

        return component_list
    #
    # def fetch_camapaigns_enviroment(strcampaignname):
    #     # Fetch campaigns and devices list against execute column status "Yes"
    #     # fetch current directory
    #     current_dir = os.getcwd()
    #     # Remove the "utilis" directory from the path
    #     parent_dir = os.path.dirname(os.path.dirname(current_dir))
    #     # create ConfigParser object
    #     config = configparser.ConfigParser()
    #     # read properties from config file
    #     ConfigProperties_Path = parent_dir + "\\RantCell_Automation\\configuration\\config.properties"
    #     config.read(ConfigProperties_Path)
    #     # get file path and file name from properties
    #     Test_data_path = config.get('Test_ExcelData', 'Test_data_path')
    #     file_name = config.get('Test_ExcelData', 'file_name')
    #     SHEET_CAMPAIGNS_TOTEST = config.get('Test_ExcelData','SHEET_CAMPAIGNS_TOTEST')
    #     SHEET_ENVIRONMENTS_USERINPUT_LOGIN = config.get('Test_ExcelData','SHEET_ENVIRONMENTS_USERINPUT_LOGIN')
    #     test_data_file_path = parent_dir + "\\" + Test_data_path + "\\" + file_name
    #     campaignwb = openpyxl.load_workbook(test_data_file_path)
    #     campaignsheet = campaignwb[SHEET_CAMPAIGNS_TOTEST]
    #     campaignsheetrows = campaignsheet.max_row
    #     campaigns_test = []
    #     for r in range(2, campaignsheetrows + 1):
    #         values = ''
    #         if campaignsheet.cell(row=r, column=3).value == "Yes" and campaignsheet.cell(row=r,column=2).value == strcampaignname:
    #             for c in range(1, 3):
    #                 value = campaignsheet.cell(row=r, column=c).value.strip()
    #                 values = values + value + ","
    #             temp = values.split(",")
    #             campaigns_test.append(tuple(temp))
    #     # print(campaigns_test)
    #
    #     # Fetch environment details to carry out execution based on user input
    #
    #     environment_test = []
    #     environmentwb = openpyxl.load_workbook(test_data_file_path)
    #     environmentsheet = environmentwb[SHEET_ENVIRONMENTS_USERINPUT_LOGIN]
    #     campaignsheetrows = environmentsheet.max_row
    #     for r in range(2, campaignsheetrows + 1):
    #         values = ''
    #         if environmentsheet.cell(row=r, column=5).value == "Yes":
    #             for c in range(1, 5):
    #                 value = environmentsheet.cell(row=r, column=c).value.strip()
    #                 values = values + value + ","
    #             temp = values.split(",")
    #             environment_test.append(tuple(temp))
    #
    #     campaign_environment = []
    #     for x in campaigns_test:
    #         va = ''
    #         for y in environment_test:
    #             va = x[0] + "," + x[1] + "," + y[0] + "," + y[1] + "," + y[2] + "," + y[3]
    #             temp = va.split(",")
    #             campaign_environment.append(tuple(temp))
    #
    #     return campaign_environment
    #
    # def fetch_components(strcampaignname):
    #
    #     # Fetch components list based on campaign name
    #     # fetch current directory
    #     current_dir = os.getcwd()
    #     # Remove the "utilis" directory from the path
    #     parent_dir = os.path.dirname(os.path.dirname(current_dir))
    #     # create ConfigParser object
    #     config = configparser.ConfigParser()
    #     # read properties from config file
    #     ConfigProperties_Path = parent_dir + "\\RantCell_Automation\\configuration\\config.properties"
    #     config.read(ConfigProperties_Path)
    #     # get file path and file name from properties
    #     Test_data_path = config.get('Test_ExcelData', 'Test_data_path')
    #     file_name = config.get('Test_ExcelData', 'file_name')
    #     SHEET_TC = config.get('Test_ExcelData', 'SHEET_TC')
    #     test_data_file_path = parent_dir + "\\" + Test_data_path + "\\" +file_name
    #     component_list = []
    #     componentwb = openpyxl.load_workbook(test_data_file_path)
    #     componentsheet = componentwb[SHEET_TC]
    #     componentsheetrows = componentsheet.max_row
    #     componentsheetcoloumns = componentsheet.max_column
    #     for r in range(2, componentsheetrows + 1):
    #         if componentsheet.cell(row=r, column=4).value.strip() == strcampaignname:
    #             for c in range(5, componentsheetcoloumns + 1):
    #                 if componentsheet.cell(row=r, column=c).value == "Yes":
    #                     value = componentsheet.cell(row=1, column=c).value.replace(" ", "")
    #                     component_list.append(value)
    #
    #     return component_list